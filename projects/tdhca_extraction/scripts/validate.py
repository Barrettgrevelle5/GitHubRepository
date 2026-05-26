#!/usr/bin/env python3
"""
validate.py — Compare agent-extracted data against ground truth aggregator.

Usage:
    python validate.py <application_number>

Example:
    python validate.py 26416

Reads from:
    - agent_output/<app_number>_extracted.xlsx  (agent's work)
    - ground_truth/*.xlsx                       (office's aggregator)

Writes to:
    - validation/<app_number>_validation.txt    (comparison report)
"""

import sys
import os
import glob
from datetime import datetime

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

# Paths relative to the tdhca_extraction project folder
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
AGENT_OUTPUT_DIR = os.path.join(PROJECT_DIR, "agent_output")
GROUND_TRUTH_DIR = os.path.join(PROJECT_DIR, "ground_truth")
VALIDATION_DIR = os.path.join(PROJECT_DIR, "validation")

# Tolerance for numeric comparisons (dollar amounts)
DOLLAR_TOLERANCE = 10          # differences <= $10 are a pass
HARD_FAIL_THRESHOLD = 1000     # differences > $1,000 are a hard fail

# Formula columns — skip these entirely, the agent doesn't populate them
FORMULA_COLUMNS = {
    20, 23, 37,
    *range(44, 53),    # 44-52
    62,
    *range(63, 83),    # 63-82
    86, 94, 96, 97, 99,
    *range(100, 141),  # 100-140
}

# Exact match columns — must be identical (no tolerance)
# Unit counts, rates, terms, identifiers
EXACT_MATCH_COLUMNS = {
    3,                  # Application Number
    12,                 # Region
    17,                 # 9% or 4%
    18,                 # LIHTC Units
    19,                 # Market Units
    24,                 # Year Constructed
    28,                 # Loan 1 Term
    29,                 # Loan 1 Amortization
    30,                 # Loan 1 Interest Rate
    32,                 # DCR
    33,                 # Loan 2 Term
    34,                 # Loan 2 Amortization
    35,                 # Loan 2 Interest Rate
}

# Text columns — fuzzy comparison (normalize whitespace, case-insensitive)
TEXT_COLUMNS = {
    4,                  # Development Name
    5,                  # Developer
    6,                  # Guarantor
    7,                  # General Contractor
    8,                  # Lender
    9,                  # Equity Partner
    10,                 # City
    11,                 # County
    13,                 # Area
    14,                 # Population
    15,                 # Set-Aside
    16,                 # Activity
    25,                 # Scattered Site
    39,                 # Source of Soft Funds
}

# All hardcoded columns the agent should populate
ALL_AGENT_COLUMNS = sorted(set(range(2, 141)) - FORMULA_COLUMNS)

# Column headers (from the aggregator row 3)
COLUMN_NAMES = {
    2: "Date of UW Report",
    3: "Application Number",
    4: "Development Name",
    5: "Developer",
    6: "Guarantor",
    7: "General Contractor",
    8: "Lender / Loan Type",
    9: "Equity Partner",
    10: "City",
    11: "County",
    12: "Region",
    13: "Area",
    14: "Population",
    15: "Set-Aside",
    16: "Activity",
    17: "9% or 4%?",
    18: "LIHTC Units",
    19: "Market Units",
    21: "Total Unit SF - NRA",
    22: "Development Site Size (Acres)",
    24: "Year Constructed",
    25: "Scattered Site?",
    26: "Construction Loan Principal",
    27: "Construction Loan Interest Rate",
    28: "Loan (1) Term",
    29: "Loan (1) Amortization",
    30: "Loan (1) Interest Rate",
    31: "Loan (1) Principal",
    32: "Debt Coverage Ratio",
    33: "Loan (2) Term",
    34: "Loan (2) Amortization",
    35: "Loan (2) Interest Rate",
    36: "Loan (2) Principal",
    38: "Total Soft Sources",
    39: "Source of Soft Funds",
    40: "Equity Price Per Credit",
    41: "Total Equity",
    42: "Income From Operations",
    43: "Deferred Developer Fee",
    53: "Acquisition Cost",
    54: "Off-Site + Site Work Cost",
    55: "Building Hard Costs",
    56: "Hard Cost Contingency",
    57: "Contractor Fee",
    58: "Soft Costs",
    59: "Financing Costs",
    60: "Developer Fee",
    61: "Reserves",
    83: "Potential Gross Rent",
    84: "Other Income",
    85: "Vacancy + Concessions",
    87: "G&A + Compliance Fees",
    88: "Management Fee",
    89: "Payroll",
    90: "Repairs & Maintenance",
    91: "Utilities",
    92: "Insurance",
    93: "Property Taxes",
    95: "Replacement Reserves",
    98: "Hard Debt Service",
}


# ---------------------------------------------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------------------------------------------

def normalize_text(val):
    """Normalize a text value for fuzzy comparison."""
    if val is None:
        return ""
    s = str(val).strip().lower()
    # collapse multiple spaces
    s = " ".join(s.split())
    # remove trailing punctuation variations
    s = s.rstrip(".")
    return s


def to_number(val):
    """Try to convert a value to float. Returns None if not numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        cleaned = str(val).replace(",", "").replace("$", "").strip()
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def format_value(val):
    """Format a value for display in the report."""
    if val is None:
        return "<empty>"
    if isinstance(val, float):
        if val == int(val) and abs(val) < 1e10:
            return f"{int(val):,}"
        if abs(val) < 1:
            return f"{val:.4f}"
        return f"{val:,.2f}"
    return str(val)


def find_ground_truth_file():
    """Find the aggregator xlsx in the ground truth folder."""
    patterns = [
        os.path.join(GROUND_TRUTH_DIR, "*.xlsx"),
        os.path.join(GROUND_TRUTH_DIR, "*.xlsb.xlsx"),
    ]
    for pattern in patterns:
        files = glob.glob(pattern)
        if files:
            return files[0]
    return None


def find_agent_output(app_number):
    """Find the agent's extracted file for a given application number."""
    path = os.path.join(AGENT_OUTPUT_DIR, f"{app_number}_extracted.xlsx")
    if os.path.exists(path):
        return path
    # Try finding any file with the app number in the name
    for f in glob.glob(os.path.join(AGENT_OUTPUT_DIR, f"*{app_number}*")):
        return f
    return None


def load_ground_truth_row(filepath, app_number):
    """
    Load the ground truth aggregator and find the row matching the
    application number. Returns a dict of {column_number: value}.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Search column 3 (Application Number) for a match
    target_row = None
    for row in range(4, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=3).value
        if cell_val is not None and str(cell_val).strip() == str(app_number).strip():
            target_row = row
            break

    if target_row is None:
        return None

    # Extract all values from that row
    data = {}
    for col in ALL_AGENT_COLUMNS:
        val = ws.cell(row=target_row, column=col).value
        # Skip formula strings — we want calculated values only
        if isinstance(val, str) and val.startswith("="):
            continue
        data[col] = val

    wb.close()
    return data


def load_agent_output(filepath):
    """
    Load the agent's extracted xlsx. Expects column headers in row 1
    matching the column numbers or names, and data in row 2.
    Returns a dict of {column_number: value}.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    data = {}
    # Read headers from row 1 to figure out which column number each maps to
    header_to_col = {}
    for excel_col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=excel_col).value
        if header is not None:
            header_str = str(header).strip()
            # Check if header is a number (column number directly)
            try:
                col_num = int(header_str)
                header_to_col[excel_col] = col_num
            except ValueError:
                # Try matching by name
                for col_num, col_name in COLUMN_NAMES.items():
                    if normalize_text(header_str) == normalize_text(col_name):
                        header_to_col[excel_col] = col_num
                        break

    # Read values from row 2
    for excel_col, col_num in header_to_col.items():
        val = ws.cell(row=2, column=excel_col).value
        data[col_num] = val

    wb.close()
    return data


# ---------------------------------------------------------------------------
# COMPARISON ENGINE
# ---------------------------------------------------------------------------

def compare(agent_data, truth_data):
    """
    Compare agent-extracted values against ground truth.

    Returns three lists:
        passes       — (col, name, agent_val, truth_val, note)
        soft_mismatches — same format, minor issues
        hard_mismatches — same format, real problems
    """
    passes = []
    soft_mismatches = []
    hard_mismatches = []

    # Only compare columns that exist in BOTH datasets
    columns_to_check = sorted(
        set(agent_data.keys()) & set(truth_data.keys()) & set(ALL_AGENT_COLUMNS)
    )

    # Also flag columns in truth but missing from agent
    missing_from_agent = sorted(
        set(truth_data.keys()) - set(agent_data.keys()) - FORMULA_COLUMNS
    )

    for col in columns_to_check:
        agent_val = agent_data[col]
        truth_val = truth_data[col]
        col_name = COLUMN_NAMES.get(col, f"Column {col}")

        # Both empty — pass
        if agent_val is None and truth_val is None:
            passes.append((col, col_name, agent_val, truth_val, "Both empty"))
            continue

        # One empty, one not — mismatch
        if agent_val is None and truth_val is not None:
            hard_mismatches.append((col, col_name, agent_val, truth_val, "Agent value missing"))
            continue
        if agent_val is not None and truth_val is None:
            soft_mismatches.append((col, col_name, agent_val, truth_val, "Extra value from agent (not in ground truth)"))
            continue

        # --- TEXT COLUMNS: fuzzy comparison ---
        if col in TEXT_COLUMNS:
            if normalize_text(agent_val) == normalize_text(truth_val):
                passes.append((col, col_name, agent_val, truth_val, "Text match"))
            else:
                soft_mismatches.append((col, col_name, agent_val, truth_val, "Text variation"))
            continue

        # --- EXACT MATCH COLUMNS: must be identical ---
        if col in EXACT_MATCH_COLUMNS:
            agent_num = to_number(agent_val)
            truth_num = to_number(truth_val)
            if agent_num is not None and truth_num is not None:
                if agent_num == truth_num:
                    passes.append((col, col_name, agent_val, truth_val, "Exact match"))
                else:
                    hard_mismatches.append((col, col_name, agent_val, truth_val,
                        f"Exact match required — off by {abs(agent_num - truth_num):.4f}"))
            elif str(agent_val).strip() == str(truth_val).strip():
                passes.append((col, col_name, agent_val, truth_val, "Exact match (string)"))
            else:
                hard_mismatches.append((col, col_name, agent_val, truth_val, "Exact match failed"))
            continue

        # --- NUMERIC COLUMNS: tolerance-based comparison ---
        agent_num = to_number(agent_val)
        truth_num = to_number(truth_val)

        if agent_num is not None and truth_num is not None:
            diff = abs(agent_num - truth_num)
            if diff == 0:
                passes.append((col, col_name, agent_val, truth_val, "Exact match"))
            elif diff <= DOLLAR_TOLERANCE:
                passes.append((col, col_name, agent_val, truth_val,
                    f"Within tolerance (diff: ${diff:,.2f})"))
            elif diff <= HARD_FAIL_THRESHOLD:
                soft_mismatches.append((col, col_name, agent_val, truth_val,
                    f"Minor difference: ${diff:,.2f}"))
            else:
                hard_mismatches.append((col, col_name, agent_val, truth_val,
                    f"SIGNIFICANT difference: ${diff:,.2f}"))
            continue

        # --- FALLBACK: string comparison ---
        if str(agent_val).strip() == str(truth_val).strip():
            passes.append((col, col_name, agent_val, truth_val, "String match"))
        else:
            soft_mismatches.append((col, col_name, agent_val, truth_val, "Value mismatch"))

    # Add missing columns as hard mismatches
    for col in missing_from_agent:
        col_name = COLUMN_NAMES.get(col, f"Column {col}")
        truth_val = truth_data[col]
        if truth_val is not None:
            hard_mismatches.append((col, col_name, None, truth_val, "Missing from agent output"))

    return passes, soft_mismatches, hard_mismatches


# ---------------------------------------------------------------------------
# REPORT GENERATOR
# ---------------------------------------------------------------------------

def generate_report(app_number, passes, soft_mismatches, hard_mismatches):
    """Generate a human-readable validation report."""
    total = len(passes) + len(soft_mismatches) + len(hard_mismatches)
    lines = []

    lines.append("=" * 70)
    lines.append(f"  VALIDATION REPORT — Application #{app_number}")
    lines.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 70)
    lines.append("")

    # --- SUMMARY ---
    lines.append("SUMMARY")
    lines.append("-" * 40)
    lines.append(f"  Fields compared:    {total}")
    lines.append(f"  Passes:             {len(passes)}")
    lines.append(f"  Soft mismatches:    {len(soft_mismatches)}  (review, probably OK)")
    lines.append(f"  Hard mismatches:    {len(hard_mismatches)}  (needs investigation)")
    lines.append("")

    if total > 0:
        accuracy = len(passes) / total * 100
        lines.append(f"  Accuracy:           {accuracy:.1f}%")
        if accuracy >= 95:
            lines.append(f"  Status:             PASS — ready to scale")
        elif accuracy >= 85:
            lines.append(f"  Status:             REVIEW — close but check mismatches")
        else:
            lines.append(f"  Status:             FAIL — too many mismatches")
    lines.append("")

    # --- HARD MISMATCHES (show first — most important) ---
    if hard_mismatches:
        lines.append("")
        lines.append("HARD MISMATCHES — Needs Investigation")
        lines.append("-" * 70)
        for col, name, agent_val, truth_val, note in hard_mismatches:
            lines.append(f"  Col {col:>3} | {name}")
            lines.append(f"          Agent:  {format_value(agent_val)}")
            lines.append(f"          Office: {format_value(truth_val)}")
            lines.append(f"          Issue:  {note}")
            lines.append("")

    # --- SOFT MISMATCHES ---
    if soft_mismatches:
        lines.append("")
        lines.append("SOFT MISMATCHES — Review (Probably OK)")
        lines.append("-" * 70)
        for col, name, agent_val, truth_val, note in soft_mismatches:
            lines.append(f"  Col {col:>3} | {name}")
            lines.append(f"          Agent:  {format_value(agent_val)}")
            lines.append(f"          Office: {format_value(truth_val)}")
            lines.append(f"          Note:   {note}")
            lines.append("")

    # --- PASSES (collapsed) ---
    lines.append("")
    lines.append(f"PASSES ({len(passes)} fields)")
    lines.append("-" * 70)
    for col, name, agent_val, truth_val, note in passes:
        lines.append(f"  Col {col:>3} | {name:<35} | {format_value(agent_val)}")

    lines.append("")
    lines.append("=" * 70)
    lines.append("  END OF REPORT")
    lines.append("=" * 70)

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python validate.py <application_number>")
        print("Example: python validate.py 26416")
        sys.exit(1)

    app_number = sys.argv[1]
    print(f"Validating application #{app_number}...")

    # Find files
    truth_file = find_ground_truth_file()
    if not truth_file:
        print(f"ERROR: No aggregator file found in {GROUND_TRUTH_DIR}")
        print("Make sure the .xlsx aggregator is in the ground_truth/ folder.")
        sys.exit(1)
    print(f"  Ground truth: {os.path.basename(truth_file)}")

    agent_file = find_agent_output(app_number)
    if not agent_file:
        print(f"ERROR: No agent output found for #{app_number}")
        print(f"Expected: {AGENT_OUTPUT_DIR}/{app_number}_extracted.xlsx")
        sys.exit(1)
    print(f"  Agent output: {os.path.basename(agent_file)}")

    # Load data
    print("  Loading ground truth...")
    truth_data = load_ground_truth_row(truth_file, app_number)
    if truth_data is None:
        print(f"ERROR: Application #{app_number} not found in aggregator.")
        print("Check that the application number is correct.")
        sys.exit(1)
    print(f"  Found {len(truth_data)} fields in ground truth")

    print("  Loading agent output...")
    agent_data = load_agent_output(agent_file)
    print(f"  Found {len(agent_data)} fields in agent output")

    # Compare
    print("  Comparing...")
    passes, soft_mismatches, hard_mismatches = compare(agent_data, truth_data)

    # Generate report
    report = generate_report(app_number, passes, soft_mismatches, hard_mismatches)

    # Save report
    os.makedirs(VALIDATION_DIR, exist_ok=True)
    report_path = os.path.join(VALIDATION_DIR, f"{app_number}_validation.txt")
    with open(report_path, "w") as f:
        f.write(report)

    # Print to console too
    print("")
    print(report)
    print(f"\nReport saved to: {report_path}")


if __name__ == "__main__":
    main()
